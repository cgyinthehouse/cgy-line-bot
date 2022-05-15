from openpyxl import load_workbook
import urllib.request as httplib  # 3.x
import json
from datetime import datetime
import aiml
import ssl
import os
ssl._create_default_https_context = ssl._create_unverified_context

def openpyxl_open_sheet(filename, sheet):
    wb = load_workbook(filename)  # 讀取檔案
    return wb[sheet]  # 打開名稱爲sheet的工作表


# 取得第x行的資料
def openpyxl_get_1row(sheet, row=2):
    list1 = [sheet.cell(row=row, column=col1).value for col1 in range(1, sheet.max_column+1)]
    return list1


# 把所有ROW 加到LIST
def openpyxl_add_all_row_to_list(sheet, list_to_add):
    for row in sheet.rows:
        if row is not None:
            list_to_add.append(row)
    return list_to_add


def openpyxl_log(userid, input_value, text, timedata, file_name='Log.xlsx'):
    wb = load_workbook(file_name)
    sheet = wb.active
    sheet.append([userid, input_value, text, timedata])
    wb.save(filename=file_name)


# 用userinput查sheet，依照col查詢cell值，回傳兩個cell value
def table_lookup_2values(sheet, user_input="", keycol=1, answer_col=2, handle_col=3):
    answer = ""
    handle = ""
    for row in range(1, sheet.max_row+1):
        key_value = sheet.cell(row=row, column=keycol).value  # 關鍵字欄位的cell
        if user_input.find(str(key_value)) > -1:  # 確認使用者輸入是否含有關鍵字
            answer = sheet.cell(row=row, column=answer_col).value  # 要回復使用者的内容
            handle = sheet.cell(row=row, column=handle_col).value  # 要執行的動作
            break  # 結束迴圈
    return answer, handle


# 用userinput查sheet，依照col查詢cell值，回傳一個cell value
def table_lookup_1value(sheet, user_input="", key_col=1, res_col=2):
    answer = ""
    for row1 in range(1, sheet.max_row+1):
        key_value = sheet.cell(row=row1, column=key_col).value  # 取得資料
        if user_input.find(str(key_value)) > -1:  # 確認使用者輸入是否含有關鍵字
            answer = sheet.cell(row=row1, column=res_col).value  # 取得資料
            break
    return answer


"""
wb = load_workbook('新竹縣美食資料.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheet = wb.active                 # 打開一個工作欄

# 方法二 依照名稱 打開工作表單   (注意：工作表單名稱，請使用英文， 中文會 出現 警告訊息)

sheetnames = wb.get_sheet_names()
print(sheetnames)
# sheet = wb.get_sheet_by_name("新竹縣美食資料")
sheet = wb.get_sheet_by_name(sheetnames[0])

#######

sheet['A1'] = 87                  # 設定資料   A1
sheet.cell(row=1, column=2).value = 'OpenPyxl Tutorial' # 設定資料 B1
wb.save("sample_file.xlsx")


"""

"""
# 取得某一筆的資料  （第二筆）  方法二
def xlsxGetRow(sheet, row1=2):
    list1 = []
    col1 = 1
    while col1 < sheet.max_column:
        x = sheet.cell(row=row1, column=col1).value  # 取得資料 A2
        print(x)
        list1.append(x)
        col1 = col1 + 1
    return list1


# 取得某一欄位的所有資料

def xlsxGetCol(sheet, col1=2):
    list1 = []
    row1 = 1
    while row1 < sheet.max_row:
        x = sheet.cell(row=row1, column=col1).value  # 取得資料 A2
        print(x)
        list1.append(x)
        row1 = row1 + 1
    return list1


def HTML_ListToHTMLTable(list1):
    x = 0
    str1 = '<table border="1">'
    for row in list1:
        y = 0
        str1 = str1 + '<tr>'
        for col in row:
            str1 = str1 + "<td>" + str(col) + "</td>"
            y = y + 1

        str1 = str1 + '<tr>'
        x = x + 1
    str1 = str1 + '</table>'
    return str1


def SQL_GEtColName(tableName):
    sql = "select column_name from INFORMATION_SCHEMA.COLUMNS where table_name='" + tableName + "'"
    cursor.execute(sql)  # 執行sql指令
    db.commit()  # 資料同步儲存
    list1 = cursor.fetchall()  # 將資料轉換成陣列
    return ListToHTMLTable(list1)


def SQL_select_ToHTMLTable(db, cursor, iSQL):
    cursor.execute(iSQL)  # 執行sql指令
    db.commit()  # 資料同步儲存

    list1 = cursor.fetchall()  # 將資料轉換成陣列
    return ListToHTMLTable(list1)


def SQL_select(db, cursor, iSQL):
    cursor.execute(iSQL)  # 執行sql指令
    db.commit()  # 資料同步儲存
    list1 = cursor.fetchall()  # 將資料轉換成陣列
    return list1


def SQL_FunSQLSelect():
    header = SQL_GEtColName("persons")
    sql = "SELECT * FROM `persons` "
    str1 = SQL_select(sql)
    return header + str1


def SQL_execute(db, cursor, iSQL):
    cursor.execute(iSQL)  # 執行sql指令
    db.commit()  # 資料同步儲存


# 把ROW 所有資料 加到LIST
def List_AddAllRowToTreeList(sheet, contacts):
    print("全部的筆數：", sheet.max_row)
    n = 2
    while n <= len(sheet):
        # list1 = mylibs.openpyxl_GetRow(sheet, row1=n)
        # list1 = openpyxl_GetRow(sheet, row1=n)
        if (list1[0] != None):
            # 表格欄位[A,B,C,D,E,F.... > 0,1,2,3,4,5....]
            str1 = (list1[0], list1[1], list1[2], list1[3], list1[4])
            contacts.append(str1)
        n = n + 1
    return contacts
"""


def line_res_text(str):
    retval = [
        {
            "type": "text",
            "text": str
        }
    ]
    return retval


# Line_回傳地圖("台北101","110台北市信義區信義路五段7號", 25.0330766,121.5609268)
def line_res_map(title, address, latitude, longitude):
    returnvalue = [
        {
            "type": "location",
            "title": title,
            "address": address,  # "110台北市信義區信義路五段7號",
            "latitude": latitude,  # 25.0330766,
            "longitude": longitude,  # 121.5609268
        }
    ]
    return returnvalue


def line_res_template(title):
    returnValue = [
        {
            "type": "template",
            "altText": "This is a buttons template",
            "template": {
                "type": "buttons",
                "thumbnailImageUrl": "https://www.unisa.edu.au/siteassets/media-centre/images/heart-coffee---shutterstock_512503885_web.jpg",
                "imageAspectRatio": "rectangle",
                "imageSize": "cover",
                "imageBackgroundColor": "#FFFFFF",
                "title": title,  # "最新促銷活動",   #
                "text": "點我看更多訊息",
                "defaultAction": {
                    "type": "uri",
                    "label": "View detail",
                    "uri": "http://www.google.com"
                },
                "actions": [
                    {
                        "type": "postback",
                        "label": "活動期間",
                        "data": "即日起至5-31"
                    },
                    {
                        "type": "uri",
                        "label": "優惠內容",
                        "uri": "好友分享日，飲品買一送一"
                    },
                    {
                        "type": "uri",
                        "label": "想知道更多",
                        "uri": "http://www.yahoo.com"
                    }
                ]
            }
        }
    ]
    return returnValue

def line_get_response_old(userInput, sheet):
    answere, handle = table_lookup_2values(sheet, userInput, keycol=1, answer_col=2, handle_col=3)
    if handle == "text" or handle == "":
        if answere != "":
            returnVal = line_res_text(answere)
            logText = answere
        else:
            #  問一下ubike
            answere2 = find_taoyuan_ubike_by_name(userInput)
            if answere2 != "":
                returnVal = line_res_text(answere2)
                logText = answere
            else:
                answere = "我不知道，請電話詢問"
                returnVal = line_res_text(answere)
                logText = answere
    elif handle == "地圖":
        title, address, latitude, longitude = answere.split(",")  # string  依照關鍵字 分成多個答案 String  split to  list
        returnVal = line_res_map(title, address, float(latitude), float(longitude))
        logText = title
    elif handle == "template":
        title = answere.split(",")  # string  依照關鍵字 分成多個答案 String  split to  list
        returnVal = line_res_template(title)
        logText = title
    return returnVal, logText


def line_get_response(userInput, sheetQA):
    answer, handle = table_lookup_2values(sheetQA, userInput, keycol=1, answer_col=2, handle_col=3)
    if handle == "text" or handle == "":
        if answer != "":
            retval = line_res_text(answer)  # retval:回傳值
            logtext = answer
        else:
            # 使用者輸入範例：我想要知道健行科技大學還有多少ubike
            # 檢查使用者輸入是否含ubike
            if userInput.find("ubike") > -1:    # str2.find('str1'): 回傳str1在str2中所處的index, -1代表查不到
                answer = find_taoyuan_ubike_by_keyword(userInput)
                if answer != "":
                    retval = line_res_text(answer)
                    logtext = answer
                else:
                    answer = "桃園ubike中查無此資料"
                    retval = line_res_text(answer)
                    logtext = answer
            else:
                answer = "I don't know what you mean."
                retval = line_res_text(answer)
                logtext = answer
    elif handle == "地圖":
        title, address, latitude, longitude = answer.split(",")
        retval = line_res_map(title, address, float(latitude), float(longitude))
        logtext = title
    elif handle == "template":
        title = answer.split(",")  # fixme
        retval = line_res_template(title)
        logtext = title
    return retval, logtext


def aiml_response(userinput, kernel):
    contents = kernel.respond(userinput)
    if "WARNING: No match found for input:" in contents or not contents:
        contents = ""
    elif 'action' in contents:
        data = json.loads(contents)
        print(data)
        if 'exe' in data:
            os.system(data["exe"])
        elif 'item' in data:
            price = 0
            meal = data["meal"]
            items = data["items"]
            tax = data["tax"]
            VISA = data["VISA"]
            if meal == ' 草莓蛋糕' or meal == '草莓蛋糕':
                price = 100 * int(items)
                if tax:
                    price *= 1.05
                if VISA:
                    price *= 1.02
            contents = "一共是" + str(price)
    return contents



def line_timestamp_formatter(timestamp):  # 將line的timestamp轉成可閲讀形式
    timestamp2 = int(timestamp) / 1000    # line的單位是毫秒，fromtimestamp函式是使用微秒
    date_time = datetime.fromtimestamp(timestamp2)
    timedata = date_time.strftime("%Y/%m/%d, %H:%M:%S")
    return timedata


def find_taoyuan_ubike_by_name(stopName):   # 當使用者輸入完全符合其中站名
    # 獲取網路下載的JSON格式的字串
    url = "http://data.tycg.gov.tw/api/v1/rest/datastore/a1b4714b-3b75-4ff8-a8f2-cc377e4eaa0f?format=json&limit=9999"
    # url="https://data.tycg.gov.tw/opendata/datalist/datasetMeta/download?id=5ca2bfc7-9ace-4719-88ae-4034b9a5a55c&rid=a1b4714b-3b75-4ff8-a8f2-cc377e4eaa0f"
    req = httplib.Request(url, data=None,
                          headers={
                              'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36"})
    reponse = httplib.urlopen(req)  # 開啟連線動作
    if reponse.code == 200:  # 當連線正常時
        contents = reponse.read()  # 讀取網頁內容
        contents = contents.decode("utf-8")  # 轉換編碼為 utf-8
    # JSON格式的字串轉成Dict型別
    data1 = json.loads(contents)
    str1 = ""
    size1 = len(data1["result"]["records"])
    for x in range(size1):
        if data1["result"]["records"][x]["sna"] == stopName:  # 當使用者輸入符合其中站名
            str1 = str1 + str(x) + "場站名稱:" + data1["result"]["records"][x]["sna"] + \
                   "  可借車位數:" + data1["result"]["records"][x]["sbi"] + \
                   "/" + data1["result"]["records"][x]["tot"] + \
                   "  地址:" + data1["result"]["records"][x]["sarea"] + data1["result"]["records"][x]["sna"] + \
                   "  lat:" + data1["result"]["records"][x]["lat"] + \
                   " lng:" + data1["result"]["records"][x]["lng"]
            break
    return str1


def find_taoyuan_ubike_by_keyword(stopName):  # 當使用者輸入内含有站名關鍵字
    # 獲取網路下載的JSON格式的字串
    url = "http://data.tycg.gov.tw/api/v1/rest/datastore/a1b4714b-3b75-4ff8-a8f2-cc377e4eaa0f?format=json&limit=9999"
    # url="https://data.tycg.gov.tw/opendata/datalist/datasetMeta/download?id=5ca2bfc7-9ace-4719-88ae-4034b9a5a55c&rid=a1b4714b-3b75-4ff8-a8f2-cc377e4eaa0f"
    req = httplib.Request(url, data=None,
                          headers={
                              'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36"})
    reponse = httplib.urlopen(req)  # 開啟連線動作
    if reponse.code == 200:  # 當連線正常時
        contents = reponse.read()  # 讀取網頁內容
        contents = contents.decode("utf-8")  # 轉換編碼為 utf-8
    # JSON格式的字串轉成Dict型別
    data1 = json.loads(contents)
    str1 = ""
    size1 = len(data1["result"]["records"])
    for x in range(size1):
        if stopName.find(data1["result"]["records"][x]["sna"]) > -1:
            str1 = str1 + str(x) + "場站名稱:" + data1["result"]["records"][x]["sna"] + \
                   "  可借車位數:" + data1["result"]["records"][x]["sbi"] + \
                   "/" + data1["result"]["records"][x]["tot"] + \
                   "  地址:" + data1["result"]["records"][x]["sarea"] + data1["result"]["records"][x]["sna"] + \
                   "  lat:" + data1["result"]["records"][x]["lat"] + \
                   " lng:" + data1["result"]["records"][x]["lng"]
            break
    return str1
