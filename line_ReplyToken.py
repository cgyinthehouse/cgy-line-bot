# 技術文章
# https://developers.line.biz/en/docs/messaging-api/sending-messages/#methods-of-sending-message
import os.path
import requests
import json
import datetime
from openpyxl import load_workbook, Workbook
import socketserver as socketserver
from http.server import SimpleHTTPRequestHandler as RequestHandler

auth_token = 'IVzkSnELK9I3QJTuhaSHPF/ymv+fkLA8X6ZkLoFJS5FRJlDGvX0z42TxLadRTGcmXoC8ysu4FxycyV/yYOE9IVsqw6OiGEpPp3N4zp4zXNKh22LFbPHtWBnjwZl8WVWePSMqQAJ5TmChyBI0yFwGgQdB04t89/1O/w1cDnyilFU='


class LineGetResponse:
    def __init__(self, file, userinput):
        self.filename = file
        self.userinput = userinput
        self.sheet = load_workbook(filename=file)['response']

    # 依照使用者輸入，查詢回應及執行動作
    def get_response(self):
        questions = {}
        for row in self.sheet.rows:
            questions[row[0].value] = [row[1].value, row[2].value]
        # fixme:購買key（同個cell多個Key）

        # 根據使用者輸入回傳查表後結果
        for key in questions.keys():  # 關鍵字有在 response sheet 裏時
            if key in self.userinput and questions[key][1] == 'text':
                msg_return = self.line_reply_format(text=questions[key][0])
                break
            elif key in self.userinput and questions[key][1] == 'template':
                msg_return = self.line_reply_format(template=True)
                break
            elif key in self.userinput and questions[key][1] == 'loc':
                msg_return = self.line_reply_format(loc=questions[key][0])
                break
        else:
            if 'ubike' in self.userinput:
                msg_return = self.line_reply_format(ubike=self.userinput)
            elif 'bus' in self.userinput:
                msg_return = self.line_reply_format(bus=self.userinput)
            else:
                msg_return = {'type': "text",
                              'text': 'Unable to recognize user\'s input'}
        return msg_return

    # 轉換成line reply 格式
    def line_reply_format(self, text='', loc='', template=False, ubike='', bus=''):
        if text:
            output = {"type": "text", "text": text}
        elif loc:
            loc = loc.split(',')
            output = {
                "type": "location",
                "title": loc[0],
                "address": loc[1],  # "110台北市信義區信義路五段7號",
                "latitude": loc[2],  # 25.0330766,
                "longitude": loc[3],  # 121.5609268
            }
        elif template:
            output = {
                "type": "template",
                "altText": "This is a buttons template",
                "template": {
                    "type": "buttons",
                    "thumbnailImageUrl": 'https://external-content.duckduckgo.com/iu/?u=http%3A%2F%2F25.media.tumblr.com%2Ftumblr_mbovzzeaEB1qldni8o1_500.jpg&f=1&nofb=1',
                    "imageAspectRatio": "rectangle",
                    "imageSize": "cover",
                    "imageBackgroundColor": "#FFFFFF",
                    "title": "Menu",
                    "text": "Please select",
                    "defaultAction": {
                        "type": "uri",
                        "label": "View detail",
                        "uri": "http://www.powenko.com/download_release/get.php"
                    },
                    "actions": [
                        {
                            "type": "uri",
                            "label": "填寫線上問卷",
                            "uri": "https://forms.gle/vdHfmWijtcBTsPNX6"
                        },
                        {
                            "type": "uri",
                            "label": "產品詳細資訊",
                            "uri": "http://www.yahoo.com"
                        }
                    ]
                }
            }
        elif ubike:
            if self.ubike_info(ubike):
                output = {'type': 'text', 'text': self.ubike_info(ubike)}
            else:
                output = {'type': 'text', 'text': '桃園ubike中查無此資料'}
        elif bus:
            binfo = self.bus_info(bus)
            if '' not in binfo:
                output = [{'type': 'text', 'text': binfo[0]},
                          {"type": "location",
                           "title": binfo[3],
                           "address": 'current location',
                           "latitude": binfo[2],
                           "longitude": binfo[1]}]
            else:
                output = {'type': 'text', 'text': '桃園公車中查無此資料'}
        return output

    def ubike_info(self, station):
        url = 'https://data.tycg.gov.tw/api/v1/rest/datastore/a1b4714b-3b75-4ff8-a8f2-cc377e4eaa0f?format=json' + '&limit=300'
        headers = {
            'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36"}
        req = requests.get(url, headers=headers)
        datas = json.loads(req.text)
        info = ''
        for x in range(len(datas["result"]["records"])):
            if station.find(datas["result"]["records"][x]["sna"]) > -1:
                info = str(
                    "中文場站名稱:" + datas["result"]["records"][x]["sna"] + '\n' +
                    "場站總停車格:" + datas["result"]["records"][x]["tot"] + '\n' +
                    "場站目前車輛數:" + datas["result"]["records"][x]["sbi"] + '\n' +
                    "地址:" + datas["result"]["records"][x]["ar"] + '\n' +
                    "場站是否暫停營運" + datas["result"]["records"][x]["act"])
                break
        return info

    def bus_info(self, bus):
        url = 'https://data.tycg.gov.tw/api/v1/rest/datastore/bf55b21a-2b7c-4ede-8048-f75420344aed?format=json' + '&limit=500'
        headers = {
            'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36"}
        req = requests.get(url, headers=headers)
        data1 = json.loads(req.text)
        size1 = len(data1["result"]["records"])
        info, longitude, latitude, busid = '','','',''
        for x in range(size1):
            if data1["result"]["records"][x]["BusID"] in bus:
                busid = data1["result"]["records"][x]["BusID"]
                info = str(
                    "車輛:" + busid + '\n' +
                    "業者代號:" + data1["result"]["records"][x]["ProviderID"] + '\n' +
                    "GPS車速:" + data1["result"]["records"][x]["Speed"] + '\n' +
                    "GPS時間:" + data1["result"]["records"][x]["DataTime"] + '\n' +
                    "路線方向(1:去程,2:回程):" + data1["result"]["records"][x]["GoBack"])
                longitude = data1["result"]["records"][x]["Longitude"]
                latitude = data1["result"]["records"][x]["Latitude"]
                break
        return info, longitude, latitude, busid


class MyHandler(RequestHandler):
    def do_POST(self):
        varLen = int(self.headers['Content-Length'])  # 取得讀取進來的網路資料長度
        if varLen > 0:
            post_data = self.rfile.read(varLen)  # 讀取傳過來的資料
            data = json.loads(post_data)  # 將json格式的字串轉爲字典格式
            print(data)
            replyToken = data['events'][0]['replyToken']  # 回郵信封(Reply Token)可以不用花錢
            userId = data['events'][0]['source']['userId']
            userInput = data['events'][0]['message']['text']
            returnType = data['events'][0]['message']['type']
            time1 = data['events'][0]['timestamp']
            time2 = str(datetime.datetime.fromtimestamp(time1 / 1000))

        if type(mmm := LineGetResponse('line_response.xlsx', userInput).get_response()) != list:
            msg = [mmm]
        else:
            msg = mmm
        message = {
            "replyToken": replyToken,
            "messages": msg
        }
        # 資料回傳 到 Line 的 https 伺服器
        hed = {'Authorization': 'Bearer ' + auth_token}
        url = 'https://api.line.me/v2/bot/message/reply'
        self.send_response(200)
        self.end_headers()
        requests.post(url, json=message, headers=hed)  # 把資料HTTP POST送出去

        # 開啓/建立 log file
        if not os.path.exists('Log.xlsx'):
            log = Workbook()
            sheet1 = log.active
            sheet1.append(['UserID', 'UserInput', 'BotReply', 'Timestamp'])
        else:
            log = load_workbook('Log.xlsx')
            sheet1 = log.active
        # 把使用者輸入及回應寫入log file
        if type(message['messages']) == dict:
            if 'text' in message['messages'].keys():
                sheet1.append([userId, userInput, message['messages']['text'], time2])
        else:
            if type(message['messages']) == list:
                try:
                    sheet1.append([userId, userInput,message['messages'][0]['text'], time2])
                except KeyError:
                    sheet1.append([userId, userInput, 'nontextreply', time2])
            else:
                sheet1.append([userId, userInput, 'nontextreply', time2])
        log.save('Log.xlsx')


socketserver.TCPServer.allow_reuse_address = True  # 可以重複使用IP
httpd = socketserver.TCPServer(('0.0.0.0', 8888), MyHandler)  # 啟動WebServer   :8888
try:
    httpd.serve_forever()  # 等待用戶使用 WebServer
except:
    print("Closing the server.")
    httpd.server_close()  # 關閉 WebServer
